import openpyxl as op
import numpy as np
import statistics as st
import matplotlib
import matplotlib.pyplot as plt

from django.shortcuts import render

from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

def simple_upload(request):
    if request.method == 'POST' and len(request.FILES) != 0:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)

        uploaded_file_url = fs.url(filename)
        headers = get_headers(myfile.name)
        rows = get_data(filename)
        central_tendency = get_central_tendency(filename)
        histogramas = get_histograma(filename)
        return render(request, 'archivos_app/index.html', {'uploaded_file_url' : uploaded_file_url, 
                                                           'headers': headers,
                                                           'rows': rows,
                                                           'central_tendency': central_tendency,
                                                           'histogramas': histogramas})
    return render(request, 'archivos_app/index.html')

def get_data(filename):
    workbook = op.load_workbook('media/'+filename)
    sheet = workbook.active
    all_rows = []
    rows = sheet.rows
    for row in rows:
        row_values = [cell.value for cell in row]
        all_rows.append(row_values)


    return all_rows[1:]


def get_headers(filename, sheetname=None):
    # Cargar el archivo de Excel
    workbook = op.load_workbook('media/'+ filename)

    # Si no se especifica una hoja, obtener la hoja activa
    if not sheetname:
        sheet = workbook.active
    else:
        sheet = workbook[sheetname]

    # Asumiendo que los encabezados están en la primera fila
    headers = [cell.value for cell in sheet[1]]

    return headers

def nosotros(request):
    return render(request, 'archivos_app/nosotros.html')

def get_central_tendency(filename):
    workbook = op.load_workbook('media/'+ filename)
    sheet = workbook.active

    resultados = {}

    for col in sheet.iter_cols(values_only=True):
        # Filtrar sólo los datos numéricos en la columna
        data = [value for value in col if isinstance(value, (int, float))]

        # Si hay datos numéricos en la columna, calcular las estadísticas
        if data:
            media = round(np.mean(data),3)
            mediana = round(np.median(data),3)
            moda = round(st.mode(data),3)
            varianza = round(st.variance(data),3)
            desviacion_estandar = round(np.sqrt(varianza),3)
            error_tipico = desviacion_estandar / np.sqrt(len(data))
            
            # Guardar los resultados en el diccionario
            resultados[col[0]] = [
                media,
                mediana,
                moda,
                varianza,
                desviacion_estandar,
                round(error_tipico,3)
                
            ]
    
    workbook.close()
    return resultados

def get_histograma(filename):
    workbook = op.load_workbook('media/'+ filename)
    sheet = workbook.active

    resultados = []

    for col in sheet.iter_cols(values_only=True):
        print(col[0])
        # Filtrar sólo los datos numéricos en la columna
        data = [value for value in col if isinstance(value, (int, float))]

        # Si hay datos numéricos en la columna, calcular las estadísticas
        if data:
            intervalos = range(int(min(data)), int(max(data)) + 4)
            plt.hist(x=data, bins=intervalos, color='#F2AB6D', rwidth=0.85)
            plt.title(f"Histograma de {col[0]}")
            plt.xlabel(col[0])
            plt.ylabel('Frecuencia')
            plt.xticks(intervalos)
            url = f"media/histograma_{col[0]}.jpg"
            plt.savefig(url)
            plt.close()
            resultados.append(url)
    
    return resultados
            

